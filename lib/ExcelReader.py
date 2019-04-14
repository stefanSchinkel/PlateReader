""" ExcelReader - for reading data from excel files
"""
# pylint: disable=C0103

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl

from .DataSet import DataSet


class ExcelReader(object):
    """Excelreader - Wrapper for SetReader

    Attributes:
        filename (str): full path of input excel file
        wb (openpyxl.Workbook): handle to workbook
        sheets (dict) : dict with sheets
    """
    def __init__(self, filename):
        """ Instantiation of a reader

        Args:
            filename (str): path to excel file
        """
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.sheets = {}

    def main(self):
        """Main wrapper"""
        for sheet in self.wb.sheetnames:
            sr = SetReader(self.wb[sheet])
            sr.main()
            self.sheets[sheet] = sr.sets


class SetReader(object):
    """SetReader - Read data from one sheet in an Excel file

    Attributes:
        sets : dict with sets as keys, contains
    """
    def __init__(self, ws):
        """ Instantiation of a reader

        Args:
            ws (openpyxl.Worksheet): handle to worksheet
        """
        self.ws = ws
        self.sets = {}

    def main(self):
        """ Wrapper to get meta data and data
        """
        self._get_sets()
        self._get_inner_dimensions()
        self._get_outer_dimensions()
        self._read_data()

    def _get_sets(self):
        """ Search the worksheet for data set and store locations
        """
        # parse b column for start of read
        sets = []
        colB = self.ws['B']  # we know its in "B"

        # find all data starts
        for cell in colB:
            if cell and "X Read #" in str(cell.value):
                sets.append([cell.row, cell.column])

        for _set in sets:
            cell = self.ws[gcl(_set[1]-1)+str(_set[0]-2)]
            data = DataSet()
            data.row = _set[0]+1
            data.col = _set[1]
            data.col_ = gcl(_set[1])

            self.sets[cell.value] = data

    def _get_inner_dimensions(self):
        """ Get the dimensions of the innter matrix
        """
        for k, v in self.sets.items():
            n_cols = []
            n_rows = []
            idx = v.row
            idy = v.col
            while idx < self.ws.max_row:
                idy_rows = gcl(idy)
                val_x = self.ws[idy_rows+str(idx)].value

                idy_cols = gcl(idy+1)
                val_y = self.ws[idy_cols+str(idx)].value
                if not val_x:
                    break
                n_rows.append(val_x)
                n_cols.append(val_y)
                idx += 1

            v.inner_x = max(n_rows)
            v.inner_y = max(n_cols)


    def _get_outer_dimensions(self):
        """ Get the number of wells, layout of the plate
        """
        for k, v in self.sets.items():
            # we stored the data start, not the header start!
            row = str(v.row -1)
            # we skip X Read/Y Read
            col = v.col + 2
            labels = []
            while col <= self.ws.max_column:
                col_name = gcl(col)
                cell = self.ws[col_name+row]
                labels.append(cell.value)
                col += 1

            # process header for outer dimension
            uniq = set([d[0] for d in labels])
            v.labels = labels
            v.n_wells = len(labels)
            v.outer_x = len(uniq)
            v.outer_y = int(len(labels)/len(uniq))


    def _read_data(self):
        """ Read the actual data to an np array
        """
        for k, v in self.sets.items():

            # pre-alloc data
            data = []
            for i in range(v.n_wells):
                data.append(np.zeros( (v.inner_x, v.inner_y) ))

            n_rows = v.inner_x * v.inner_y
            for iRow in range(v.row, v.row+n_rows):
                # np starts at 0, PlateReader at 1
                inner_x = self.ws[gcl(v.col)+str(iRow)].value -1
                inner_y = self.ws[gcl(v.col+1)+str(iRow)].value -1
                idx_outer = 0
                for iCol in range(v.col+2, self.ws.max_column+1):
                    cell = self.ws[gcl(iCol)+str(iRow)]
                    try:
                        val = int(cell.value)
                    except (TypeError, ValueError):
                        val = 0

                    # we have to swap x and y for numpy !!!
                    data[idx_outer][(inner_y,inner_x)] = val
                    idx_outer += 1

                    v.data = data
